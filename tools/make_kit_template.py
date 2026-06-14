"""Generate the BOQ inverter-kit input template (bilingual TH/EN)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F6F54")     # green header
SUB_FILL = PatternFill("solid", fgColor="E8F3EE")
EX_FILL = PatternFill("solid", fgColor="FFF7E6")        # example rows (amber tint)
INPUT_FILL = PatternFill("solid", fgColor="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="0E3D2C")
thin = Side(style="thin", color="C9D6CF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hcell(ws, r, c, val):
    cell = ws.cell(r, c, val)
    cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    cell.fill = HEAD_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    return cell

def dcell(ws, r, c, val, fill=None, bold=False, color="000000", align="left"):
    cell = ws.cell(r, c, val)
    cell.font = Font(name=FONT, size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER
    if fill:
        cell.fill = fill
    return cell

wb = Workbook()

# ── Lists sheet (dropdown sources) ───────────────────────────────────────────
ls = wb.active
ls.title = "Lists"
LISTS = {
    "A": ("System", ["micro", "string", "hybrid"]),
    "B": ("Phase", ["1", "3"]),
    "C": ("KitPhase", ["any", "1", "3"]),
    "D": ("Scaling", ["per_job", "per_inverter", "per_string", "per_panel",
                       "per_battery_module", "per_battery_stack"]),
    "E": ("Condition", ["always", "if_battery", "if_backup_smartguard",
                         "if_backup_backupbox", "if_optimizer", "if_extra_panel"]),
    "F": ("Group", ["INVERTER", "COMBINER BOX", "BATTERY", "BACKUP", "CABLE", "ACCESSORIES"]),
    "G": ("Unit", ["ตัว", "ชุด", "เส้น", "อัน", "ก้อน", "ตู้", "ชิ้น", "LOT", "SET", "ม."]),
}
for col, (name, vals) in LISTS.items():
    ls[f"{col}1"] = name
    ls[f"{col}1"].font = Font(name=FONT, bold=True)
    for i, v in enumerate(vals, start=2):
        ls[f"{col}{i}"] = v
        ls[f"{col}{i}"].font = Font(name=FONT, size=10)
    ls.column_dimensions[col].width = 20
ls.sheet_state = "hidden"

def listref(col, n):
    return f"Lists!${col}$2:${col}${n+1}"

# ── Instructions sheet ───────────────────────────────────────────────────────
ins = wb.create_sheet("📖 วิธีใช้ Instructions")
ins.sheet_view.showGridLines = False
ins.column_dimensions["A"].width = 3
ins.column_dimensions["B"].width = 116
def line(r, text, bold=False, size=10, color="000000", fill=None):
    c = ins.cell(r, 2, text)
    c.font = Font(name=FONT, bold=bold, size=size, color=color)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    if fill:
        c.fill = fill
    return c
line(1, "  แบบฟอร์มกรอกชุดอุปกรณ์อินเวอร์เตอร์ → ใช้สร้าง BOQ อัตโนมัติ", True, 14, "FFFFFF", TITLE_FILL)
ins.row_dimensions[1].height = 30
rows = [
    ("", 6),
    ("ฟอร์มนี้แบ่งเป็น 2 ระบบหลักตามที่ตกลงกัน — กรอกแล้วส่งกลับมา ระบบจะนำไปอัปเดตฐานข้อมูลและคำนวณ BOQ ให้อัตโนมัติ", True, 11, "0E3D2C"),
    ("", 4),
    ("ชีตในไฟล์นี้ / Sheets:", True, 11, "1F6F54"),
    ("   1) ‘1·อินเวอร์เตอร์ Inverters’ — รายการรุ่นอินเวอร์เตอร์ + สเปค (ใช้คำนวณจำนวนตัว/ขนาด RCBO)", False, 10),
    ("   2) ‘2·ชุดไมโคร Micro Kit’ — อุปกรณ์ของระบบ Microinverter (ATMOCE)", False, 10),
    ("   3) ‘3·ชุดสตริง String Kit’ — อุปกรณ์ของระบบ String/Hybrid (Huawei ฯลฯ)", False, 10),
    ("", 6),
    ("กติกาสำคัญ / Key rules:", True, 11, "1F6F54"),
    ("   • ‘ชื่ออุปกรณ์’ ต้องสะกดให้ตรงกับชื่อในคลังสินค้า เพื่อให้ดึงราคาอัตโนมัติได้ (แถวตัวอย่างพื้นเหลืองคือชื่อที่ตรงอยู่แล้ว)", False, 10),
    ("   • คอลัมน์ ‘รุ่น Model’: ใส่ชื่อรุ่นเจาะจง หรือพิมพ์ ‘ALL’ = ใช้กับทุกรุ่นในระบบนั้น", False, 10),
    ("   • คอลัมน์ ‘คิดตาม Scaling’ บอกวิธีคูณจำนวน (ดูตารางด้านล่าง)", False, 10),
    ("   • คอลัมน์ ‘เฟส Phase’: any = ใช้ได้ทุกเฟส, 1 = เฉพาะ 1 เฟส, 3 = เฉพาะ 3 เฟส (ระบบจะเลือกตามงาน)", False, 10),
    ("   • คอลัมน์ ‘เงื่อนไข Condition’: ใส่อุปกรณ์นั้นเฉพาะเมื่อเข้าเงื่อนไข เช่น if_battery = ใส่เมื่อมีแบต", False, 10),
    ("   • แถวพื้นเหลือง = ตัวอย่างจากระบบปัจจุบัน (แก้ได้) · แถวว่างด้านล่าง = พื้นที่กรอกเพิ่ม", False, 10),
    ("", 6),
    ("ความหมายของ ‘คิดตาม Scaling’:", True, 11, "1F6F54"),
]
r = 2
for item in rows:
    if len(item) == 2:
        ins.row_dimensions[r].height = item[1]
        r += 1
        continue
    text, *rest = item
    bold = rest[0] if len(rest) > 0 else False
    size = rest[1] if len(rest) > 1 else 10
    color = rest[2] if len(rest) > 2 else "000000"
    line(r, text, bold, size, color)
    r += 1

scaling_doc = [
    ("per_job", "1 ชุดต่อ 1 งาน (× จำนวนในช่อง ‘จำนวน/หน่วย’)", "ตู้ Combiner, Smart Meter, Dongle, Backup"),
    ("per_inverter", "× จำนวนอินเวอร์เตอร์", "ตัวอินเวอร์เตอร์, AC SPD, RCBO"),
    ("per_string", "× จำนวน String รวม (ใส่ 2 ในช่องจำนวน = String×2)", "DC Fuse, Fuse Holder, MC4"),
    ("per_panel", "× จำนวนแผง", "Optimizer (1:1 ต่อแผง)"),
    ("per_battery_module", "× จำนวนก้อนแบต = ปัดขึ้น(kWh ÷ 7)", "LUNA2000-S1"),
    ("per_battery_stack", "× จำนวนแสตก = ปัดขึ้น(ก้อนแบต ÷ 3)", "LUNA2000-C1 Power Module"),
]
hr = r
for j, t in enumerate(["คิดตาม Scaling", "ความหมาย", "ตัวอย่างอุปกรณ์"]):
    hcell(ins, hr, 2 + j, t)
ins.row_dimensions[hr].height = 22
# widen helper cols
ins.column_dimensions["C"].width = 52
ins.column_dimensions["D"].width = 40
for k, (a, b, c) in enumerate(scaling_doc, start=hr + 1):
    dcell(ins, k, 2, a, bold=True, color="1F6F54")
    dcell(ins, k, 3, b)
    dcell(ins, k, 4, c)
note_r = hr + len(scaling_doc) + 2
line(note_r, "   หมายเหตุ: RCBO ขนาดคิดอัตโนมัติจาก ‘กระแสออก × 1.25’ ปัดขึ้นขนาดมาตรฐาน — ไม่ต้องระบุแอมป์เอง", False, 10, "B45309")

# ── Sheet 1: Inverters ───────────────────────────────────────────────────────
n_sys = len(LISTS["A"][1]); n_ph = len(LISTS["B"][1])
inv = wb.create_sheet("1·อินเวอร์เตอร์ Inverters")
inv.sheet_view.showGridLines = False
inv_headers = [
    ("ระบบ\nSystem", 12), ("แบรนด์\nBrand", 14), ("รุ่น (ตรงกับคลัง)\nModel", 38),
    ("kW", 8), ("เฟส\nPhase", 8), ("แผงต่อตัว\nPanels/inv\n(micro)", 12),
    ("MAX PV (kW)\n(string/hybrid)", 14), ("จำนวน String\nString inputs", 13),
    ("กระแสออก A\nOutput A\n(→ RCBO)", 12), ("หน่วย\nUnit", 9), ("หมายเหตุ\nNote", 26),
]
for c, (t, w) in enumerate(inv_headers, 1):
    hcell(inv, 1, c, t)
    inv.column_dimensions[inv.cell(1, c).column_letter].width = w
inv.row_dimensions[1].height = 44
inv.freeze_panes = "A2"

inv_examples = [
    ["micro", "ATMOCE", "ATMOCE Micro-inverter 500Watt 1:1", 0.5, "1", 1, "", "", "", "ตัว", "อัตรา 1:1 (1 แผง/ตัว)"],
    ["micro", "ATMOCE", "ATMOCE Micro-inverter 1250Watt 2:1 ", 1.25, "1", 2, "", "", "", "ตัว", "อัตรา 2:1 (2 แผง/ตัว)"],
    ["hybrid", "Huawei", "Huawei SUN2000-5K-LB0", 5, "1", "", 7.5, 2, 22.7, "ตัว", "ตัวอย่าง — โปรดตรวจสเปคจริง"],
    ["string", "Huawei", "Huawei SUN2000-10K-LC0", 10, "1", "", 15, 2, 45.5, "ตัว", "ตัวอย่าง — โปรดตรวจสเปคจริง"],
    ["string", "Huawei", "Huawei SUN2000-10KTL-M1", 10, "3", "", 15, 2, 16.9, "ตัว", "ตัวอย่าง — โปรดตรวจสเปคจริง"],
]
row = 2
for ex in inv_examples:
    for c, v in enumerate(ex, 1):
        dcell(inv, row, c, v, fill=EX_FILL, align="center" if c not in (3, 11) else "left")
    row += 1
INV_LAST = row + 30
for rr in range(row, INV_LAST):
    for c in range(1, len(inv_headers) + 1):
        dcell(inv, rr, c, None, fill=INPUT_FILL, align="center" if c not in (3, 11) else "left")

dv_sys = DataValidation(type="list", formula1=listref("A", n_sys), allow_blank=True); inv.add_data_validation(dv_sys)
dv_ph = DataValidation(type="list", formula1=listref("B", n_ph), allow_blank=True); inv.add_data_validation(dv_ph)
dv_unit_i = DataValidation(type="list", formula1=listref("G", len(LISTS["G"][1])), allow_blank=True); inv.add_data_validation(dv_unit_i)
dv_sys.add(f"A2:A{INV_LAST-1}")
dv_ph.add(f"E2:E{INV_LAST-1}")
dv_unit_i.add(f"J2:J{INV_LAST-1}")

# ── Kit sheets (Micro + String) ──────────────────────────────────────────────
kit_headers = [
    ("รุ่น Model\n(หรือ ALL)", 30), ("กลุ่ม BOQ\nGroup", 15), ("ชื่ออุปกรณ์ (ตรงกับคลัง)\nEquipment Name", 42),
    ("หน่วย\nUnit", 9), ("จำนวน/หน่วย\nQty per unit", 11), ("คิดตาม\nScaling", 18),
    ("เฟส\nPhase", 9), ("เงื่อนไข\nCondition", 20), ("หมายเหตุ\nNote", 26),
]
nD = len(LISTS["D"][1]); nE = len(LISTS["E"][1]); nF = len(LISTS["F"][1]); nG = len(LISTS["G"][1]); nC = len(LISTS["C"][1])

def build_kit(title, examples):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    for c, (t, w) in enumerate(kit_headers, 1):
        hcell(ws, 1, c, t)
        ws.column_dimensions[ws.cell(1, c).column_letter].width = w
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"
    row = 2
    for ex in examples:
        for c, v in enumerate(ex, 1):
            dcell(ws, row, c, v, fill=EX_FILL, align="center" if c in (4, 5, 6, 7) else "left")
        row += 1
    last = row + 30
    for rr in range(row, last):
        for c in range(1, len(kit_headers) + 1):
            dcell(ws, rr, c, None, fill=INPUT_FILL, align="center" if c in (4, 5, 6, 7) else "left")
    dvU = DataValidation(type="list", formula1=listref("G", nG), allow_blank=True); ws.add_data_validation(dvU)
    dvS = DataValidation(type="list", formula1=listref("D", nD), allow_blank=True); ws.add_data_validation(dvS)
    dvP = DataValidation(type="list", formula1=listref("C", nC), allow_blank=True); ws.add_data_validation(dvP)
    dvC = DataValidation(type="list", formula1=listref("E", nE), allow_blank=True); ws.add_data_validation(dvC)
    dvG = DataValidation(type="list", formula1=listref("F", nF), allow_blank=True); ws.add_data_validation(dvG)
    dvG.add(f"B2:B{last-1}")
    dvU.add(f"D2:D{last-1}")
    dvS.add(f"F2:F{last-1}")
    dvP.add(f"G2:G{last-1}")
    dvC.add(f"H2:H{last-1}")
    return ws

# Micro kit — from current ATMOCE logic
micro_ex = [
    ["ALL", "INVERTER", "(ตัวไมโครคิดอัตโนมัติ = แผง ÷ อัตราต่อตัว)", "LOT", 1, "per_inverter", "any", "always", "ไม่ต้องกรอก ระบบใส่ให้"],
    ["ALL", "COMBINER BOX", "M-Combiner 1P (MC-100)", "ชุด", 1, "per_job", "1", "always", ""],
    ["ALL", "COMBINER BOX", "M-Combiner 3P (MC-100T)", "ชุด", 1, "per_job", "3", "always", ""],
    ["ALL", "COMBINER BOX", "CT 250A x1", "ชุด", 1, "per_job", "1", "always", ""],
    ["ALL", "COMBINER BOX", "CT 250A x3", "ชุด", 1, "per_job", "3", "always", ""],
    ["ALL", "BACKUP", "M-Backup 1P (MU100-S)", "ชุด", 1, "per_job", "1", "if_backup_backupbox", ""],
    ["ALL", "BACKUP", "M-Backup 3P (MU100-T)", "ชุด", 1, "per_job", "3", "if_backup_backupbox", ""],
    ["ALL", "BATTERY", "7kWh M-Battery (MS-7K-U)", "ก้อน", 1, "per_battery_module", "any", "if_battery", "1 ก้อน/7kWh"],
    ["ALL", "CABLE", "Single-phase junction adapter", "ชุด", 1, "per_job", "1", "always", ""],
    ["ALL", "CABLE", "Three-phase junction adapter", "ชุด", 1, "per_job", "3", "always", ""],
    ["ALL", "CABLE", "1.3 m, Three-terminal AC Cable (MW-025013-A)", "ชุด", 1, "per_inverter", "any", "always", ""],
    ["ALL", "CABLE", "2 m, Two-terminal AC Cable (MW-025020-B0)", "ชุด", 1, "per_inverter", "any", "always", "ระบบหัก 3 ตัวแรก"],
]
build_kit("2·ชุดไมโคร Micro Kit", micro_ex)

# String kit — from current Huawei logic
string_ex = [
    ["ALL", "INVERTER", "(ตัวอินเวอร์เตอร์คิดอัตโนมัติจาก MAX PV)", "ตัว", 1, "per_inverter", "any", "always", "ไม่ต้องกรอก ระบบใส่ให้"],
    ["ALL", "INVERTER", "Smart Meter DDSU666-H + CT 100A/40mA (1 เฟส)", "ชุด", 1, "per_job", "1", "always", ""],
    ["ALL", "INVERTER", "Smart Meter DTSU666-H + CT 100A/40mA (3 เฟส)", "ชุด", 1, "per_job", "3", "always", ""],
    ["ALL", "INVERTER", "Smart Dongle-WLAN-FE", "ชุด", 1, "per_job", "any", "always", "master พ่วง RS485"],
    ["ALL", "BATTERY", "HUAWEI LUNA2000-10KW-C1 (Power Module)", "ตัว", 1, "per_battery_stack", "any", "if_battery", "1/แสตก (≤3 ก้อน)"],
    ["ALL", "BATTERY", "HUAWEI LUNA2000-S1 (7kWh)", "ก้อน", 1, "per_battery_module", "any", "if_battery", "1 ก้อน/7kWh"],
    ["ALL", "BACKUP", "SmartGuard-63A-S0 (1 เฟส)", "ตัว", 1, "per_job", "1", "if_backup_smartguard", ""],
    ["ALL", "BACKUP", "SmartGuard-63A-T0 (3 เฟส)", "ตัว", 1, "per_job", "3", "if_backup_smartguard", ""],
    ["ALL", "BACKUP", "Backup Box-B0 (1 เฟส)", "ตัว", 1, "per_job", "1", "if_backup_backupbox", ""],
    ["ALL", "BACKUP", "Backup Box-B1 (3 เฟส)", "ตัว", 1, "per_job", "3", "if_backup_backupbox", ""],
    ["ALL", "INVERTER", "Smart PV Optimizer SUN2000-600W-P", "ตัว", 1, "per_panel", "any", "if_optimizer", "1:1 ต่อแผง"],
    ["ALL", "COMBINER BOX", "AC/DC Combiner Box ตู้หน้ากระจก เบอร์5", "ตู้", 1, "per_job", "any", "always", ""],
    ["ALL", "COMBINER BOX", "DC FUSE HOLDER", "ตัว", 2, "per_string", "any", "always", "String × 2"],
    ["ALL", "COMBINER BOX", "DC FUSE 16A 1000VDC", "ตัว", 2, "per_string", "any", "always", "String × 2"],
    ["ALL", "COMBINER BOX", "DC SPD 2P 800VDC 20-40KA", "ตัว", 1, "per_string", "any", "always", ""],
    ["ALL", "COMBINER BOX", "DC MCB 20A 2P 800VDC", "ตัว", 1, "per_string", "any", "always", ""],
    ["ALL", "COMBINER BOX", "MC4", "ชุด", 1, "per_string", "any", "always", ""],
    ["ALL", "COMBINER BOX", "AC SPD 2P Uc275V In20Ka/Imax40Ka", "ตัว", 1, "per_inverter", "1", "always", ""],
    ["ALL", "COMBINER BOX", "AC SPD 4P Uc385V In20Ka/Imax40Ka", "ตัว", 1, "per_inverter", "3", "always", ""],
    ["ALL", "COMBINER BOX", "RCBO (auto ตามกระแสออก×1.25)", "ตัว", 1, "per_inverter", "any", "always", "ระบบคิดขนาดให้"],
    ["ALL", "COMBINER BOX", "WIRE DUCT 40x40mm (ยาว 2 ม.)", "เส้น", 1, "per_job", "any", "always", ""],
    ["ALL", "COMBINER BOX", "DIN RAIL DNR274", "เส้น", 1, "per_job", "any", "always", ""],
    ["ALL", "COMBINER BOX", "Stopper เหล็ก รางปีกนก 2 น็อตคู่", "ตัว", 10, "per_job", "any", "always", "10/งาน"],
    ["ALL", "COMBINER BOX", "Grounding Bus-Bar 8 Slots hole 6mm", "อัน", 1, "per_job", "any", "always", ""],
]
build_kit("3·ชุดสตริง String Kit", string_ex)

wb.save("D:/Woking space/Claude code/BOQ_Inverter_Kit_Template.xlsx")
print("saved")
